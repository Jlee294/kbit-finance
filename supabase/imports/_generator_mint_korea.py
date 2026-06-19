import uuid
from openpyxl import load_workbook
from datetime import datetime

WB = "/Users/chaule/Downloads/Mint Korea.xlsx"
OUT = "/Users/chaule/Library/CloudStorage/OneDrive-Personal/KBIT HOLDINGS/Finance System/kbit/supabase/imports/mint_korea_2026.sql"
wb = load_workbook(WB, data_only=True)

def U(): return str(uuid.uuid4())
def q(s):
    if s is None: return 'null'
    return "'" + str(s).replace("'", "''").strip() + "'"
def num(x):
    if x is None or x == '': return '0'
    try: return repr(round(float(x), 4))
    except: return '0'
def d(x):
    if x is None or x == '': return 'null'
    if isinstance(x, datetime): return "'" + x.strftime('%Y-%m-%d') + "'"
    s = str(x)[:10]
    return "'" + s + "'"

def datarows(sn, start_idx):
    ws = wb[sn]
    rows = list(ws.iter_rows(values_only=True))
    return rows[start_idx:]

# ── Parse sheets ──
# Nhatkybanra data từ index 4 (dòng 5). cols: 0 Tháng,1 SốHĐ,2 NgàyHĐ,3 Mã,4 Tên,5 ĐVT,6 SL,7 ĐơnGiá,8 ThànhTiền,9 ĐịnhKhoản,10 SốPX,11 TênKH
nkb = [r for r in datarows('Nhatkybanra',4) if r[1] not in (None,'') and str(r[0]).strip() not in ('','[1]')]
# Nhatkymuavao data index 5: same layout, 10=ghichu
nkm = [r for r in datarows('Nhatkymuavao',5) if r[1] not in (None,'') and str(r[0]).strip() not in ('','[1]')]
# Bangkebanra index 5: 0 Loại,1 mẫuHĐ,2 KýhiệuHĐ,3 SốHĐ,4 NgàyHĐ,5 TênKH,6 MST,7 NộiDung,8 ThànhTiền,9 ThuếSuất,10 VAT,11 GhiChú,12 TổngHĐ
bkb = [r for r in datarows('Bangkebanra',5) if r[3] not in (None,'') and str(r[0]).strip() not in ('','[1]')]
# Bangkemuavao index 5: 0 Loại,1 mẫu,2 KH,3 SốHĐ,4 Ngày,5 TênNB,6 MST,7 MặtHàng,8 ThànhTiền,9 ThuếSuất,10 VAT,11 GC,12 TổngHĐ
bkm = [r for r in datarows('Bangkemuavao',5) if r[3] not in (None,'') and str(r[0]).strip() not in ('','[1]')]
# NXT data index 7: 1 Mã,2 Tên,3 ĐVT,4 tồnĐầuSL,5 tồnĐầuGT, 18 tồnCuốiSL, 20 giá
nxt = [r for r in datarows('NXT',7) if r[1] and str(r[0]).strip() not in ('','Stt','[1]')]

# ── IDs ──
CO = U(); WH = U()
COMPANY_NAME = 'CÔNG TY TNHH MINT KOREA'; MST='0316593003'

# Products từ NXT (master đầy đủ)
prod = {}  # code -> {id,name,unit,open_qty,open_val,close_qty,avg}
for r in nxt:
    code = str(r[1]).strip()
    if not code or code in prod: continue
    prod[code] = dict(id=U(), name=str(r[2]).strip() if r[2] else code,
                      unit=str(r[3]).strip() if r[3] else 'cái',
                      open_qty=float(r[4] or 0), open_val=float(r[5] or 0),
                      close_qty=float(r[18] or 0), avg=float(r[20] or 0))
# Bổ sung mã chỉ có trong journal (nếu thiếu)
for r in nkb+nkm:
    code = str(r[3]).strip() if r[3] else None
    if code and code not in prod:
        prod[code] = dict(id=U(), name=str(r[4]).strip() if r[4] else code,
                          unit=str(r[5]).strip() if r[5] else 'cái',
                          open_qty=0, open_val=0, close_qty=0, avg=float(r[7] or 0))

# Customers từ Bangkebanra
cust = {}  # name -> {id,code,mst}
ci=0
for r in bkb:
    name = str(r[5]).strip() if r[5] else None
    if name and name not in cust:
        ci+=1; cust[name]=dict(id=U(), code=f'KH-{ci:03d}', mst=str(r[6]).strip() if r[6] else None)
# Suppliers từ Bangkemuavao
sup={}; si=0
for r in bkm:
    name=str(r[5]).strip() if r[5] else None
    if name and name not in sup:
        si+=1; sup[name]=dict(id=U(), code=f'NCC-{si:03d}', mst=str(r[6]).strip() if r[6] else None)

out=[]
out.append("-- ════════════════════════════════════════════════════════════")
out.append("-- IMPORT DỮ LIỆU: CÔNG TY TNHH MINT KOREA (MST 0316593003) — 2026")
out.append("-- Sinh tự động từ file Excel. Chạy 1 lần trong Supabase SQL Editor.")
out.append(f"-- Sản phẩm:{len(prod)}  KH:{len(cust)}  NCC:{len(sup)}  HĐ bán:{len(bkb)}  HĐ mua:{len(bkm)}")
out.append("-- ════════════════════════════════════════════════════════════")
out.append("begin;")
out.append("set local session_replication_role = 'replica';  -- bỏ qua trigger khóa kỳ/giá vốn")
out.append("")
# Company
out.append(f"insert into companies (id, code, name, country, base_currency) values ({q(CO)}, 'MINTKOREA', {q(COMPANY_NAME)}, 'VN', 'VND');")
out.append(f"insert into warehouses (id, code, name, company_id, is_default) values ({q(WH)}, 'KHO-MK', 'Kho chính', {q(CO)}, true);")
out.append("")
# Products
out.append("-- ── Sản phẩm ──")
for code,p in prod.items():
    out.append(f"insert into products (id, code, name, unit) values ({q(p['id'])}, {q(code)}, {q(p['name'])}, {q(p['unit'])});")
out.append("")
# Customers
out.append("-- ── Khách hàng ──")
for name,c in cust.items():
    out.append(f"insert into customers (id, code, name, tax_code) values ({q(c['id'])}, {q(c['code'])}, {q(name)}, {q(c['mst'])});")
out.append("")
# Suppliers
out.append("-- ── Nhà cung cấp ──")
for name,s in sup.items():
    out.append(f"insert into suppliers (id, code, name, country, tax_code) values ({q(s['id'])}, {q(s['code'])}, {q(name)}, 'VN', {q(s['mst'])});")
out.append("")
# Opening stock + cache + warehouse_stock
out.append("-- ── Tồn đầu kỳ + giá vốn + tồn hiện tại ──")
for code,p in prod.items():
    if p['open_qty']>0:
        uc = round(p['open_val']/p['open_qty'],2) if p['open_qty'] else 0
        out.append(f"insert into warehouse_transactions (txn_type, company_id, warehouse_id, product_id, qty, txn_date, note, unit_cost, has_invoice) values ('opening', {q(CO)}, {q(WH)}, {q(p['id'])}, {num(p['open_qty'])}, '2026-01-01', 'Tồn đầu kỳ', {num(uc)}, true);")
    if p['close_qty']!=0 or p['avg']!=0:
        out.append(f"insert into product_moving_cost (company_id, product_id, qty_on_hand, avg_cost) values ({q(CO)}, {q(p['id'])}, {num(p['close_qty'])}, {num(p['avg'])});")
    if p['close_qty']>0:
        out.append(f"insert into warehouse_stock (warehouse_id, product_id, qty_on_hand) values ({q(WH)}, {q(p['id'])}, {num(p['close_qty'])});")
out.append("")

# ── SALES ORDERS ──
out.append("-- ── Đơn bán ra (customer_orders) ──")
# group sales lines by Số HĐ
nkb_by = {}
for r in nkb:
    so = str(r[1]).strip()
    nkb_by.setdefault(so, []).append(r)
bn=0
for r in bkb:
    bn+=1
    so = str(r[3]).strip(); name = str(r[5]).strip() if r[5] else None
    c = cust.get(name)
    if not c: continue
    oid=U()
    total=num(r[12]); vat=num(r[10])
    vatpct = round(float(r[9] or 0)*100,2)
    code=f'MK-BR-{bn:03d}'
    out.append(f"insert into customer_orders (id, company_id, customer_id, order_code, order_date, grand_total, amount_paid, fulfillment_status, payment_status, vat_pct, vat_amount, invoice_template, invoice_symbol, invoice_no, invoice_date, customer_tax_code, warehouse_id, stock_deducted) values ({q(oid)}, {q(CO)}, {q(c['id'])}, {q(code)}, {d(r[4])}, {total}, 0, 'delivered', 'unpaid', {vatpct}, {vat}, {q(r[1])}, {q(r[2])}, {q(r[3])}, {d(r[4])}, {q(r[6])}, {q(WH)}, true);")
    for ln in nkb_by.get(so, []):
        pc = str(ln[3]).strip() if ln[3] else None
        p = prod.get(pc)
        pid = q(p['id']) if p else 'null'
        avg = num(p['avg']) if p else '0'
        out.append(f"insert into customer_order_items (order_id, product_id, description, qty, unit_price, cost_price) values ({q(oid)}, {pid}, {q(ln[4])}, {num(ln[6])}, {num(ln[7])}, {avg});")
        # stock issue
        if p:
            out.append(f"insert into warehouse_transactions (txn_type, company_id, warehouse_id, product_id, qty, txn_date, note, unit_cost, has_invoice, ref_order_id) values ('order_deduction', {q(CO)}, {q(WH)}, {pid}, {num(ln[6])}, {d(r[4])}, {q('Bán: '+code)}, {avg}, true, {q(oid)});")
out.append("")

# ── PURCHASE ORDERS ──
out.append("-- ── Đơn mua vào (supplier_orders) ──")
nkm_by={}
for r in nkm:
    nkm_by.setdefault((str(r[1]).strip(), d(r[2])), []).append(r)
mn=0
for r in bkm:
    mn+=1
    name=str(r[5]).strip() if r[5] else None
    s=sup.get(name)
    if not s: continue
    oid=U()
    goods=num(r[8]); vat=num(r[10])
    code=f'MK-MV-{mn:03d}'
    out.append(f"insert into supplier_orders (id, company_id, supplier_id, order_code, order_type, order_date, currency, goods_value, vat_import, amount_paid, invoice_template, invoice_symbol, invoice_no, invoice_date, supplier_tax_code, vat_amount, warehouse_id, stock_added) values ({q(oid)}, {q(CO)}, {q(s['id'])}, {q(code)}, 'domestic', {d(r[4])}, 'VND', {goods}, {vat}, 0, {q(r[1])}, {q(r[2])}, {q(r[3])}, {d(r[4])}, {q(r[6])}, {vat}, {q(WH)}, false);")
    goods_lines = nkm_by.get((str(r[3]).strip(), d(r[4])), [])
    if goods_lines:
        for ln in goods_lines:
            pc=str(ln[3]).strip() if ln[3] else None
            p=prod.get(pc); pid=q(p['id']) if p else 'null'
            out.append(f"insert into supplier_order_items (order_id, product_id, description, qty, unit_price, unit_cost) values ({q(oid)}, {pid}, {q(ln[4])}, {num(ln[6])}, {num(ln[7])}, {num(ln[7])});")
            if p:
                out.append(f"insert into warehouse_transactions (txn_type, company_id, warehouse_id, product_id, qty, txn_date, note, unit_cost, has_invoice) values ('receipt', {q(CO)}, {q(WH)}, {pid}, {num(ln[6])}, {d(r[4])}, {q('Mua: '+code)}, {num(ln[7])}, true);")
    else:
        # dịch vụ: 1 dòng mô tả
        out.append(f"insert into supplier_order_items (order_id, product_id, description, qty, unit_price, unit_cost) values ({q(oid)}, null, {q(r[7])}, 1, {goods}, 0);")
out.append("")
out.append("set local session_replication_role = 'origin';")
out.append("commit;")
out.append("")
out.append("-- ── Verify ──")
out.append("select 'products' t, count(*) from products union all select 'customers', count(*) from customers union all select 'suppliers', count(*) from suppliers union all select 'customer_orders', count(*) from customer_orders union all select 'supplier_orders', count(*) from supplier_orders union all select 'warehouse_transactions', count(*) from warehouse_transactions;")

open(OUT,'w').write('\n'.join(out))
print("Đã sinh:", OUT)
print("Dòng SQL:", len(out))
print(f"company=1 warehouse=1 products={len(prod)} customers={len(cust)} suppliers={len(sup)} sales_orders={len(bkb)} purchase_orders={len(bkm)}")
